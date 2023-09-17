import openpyxl

def get_row(path,sheetname):
    workbook = openpyxl.load_workbook(path)
    sheet =workbook[sheetname]
    return sheet.max_row

def get_column(path,sheetname):
    workbook = openpyxl.load_workbook(path)
    sheet =workbook[sheetname]
    return sheet.max_column

def read_excel(path,sheetname,rowN,colN):
    workbook = openpyxl.load_workbook(path)
    sheet =workbook[sheetname]
    return sheet.cell(row=rowN,column=colN).value

def write_excel(path,sheetname,rowN,colN,data):
    workbook = openpyxl.load_workbook(path)
    sheet =workbook[sheetname]
    sheet.cell(row=rowN,column=colN).value =data
    workbook.save(path)

